"""
ClearMetric Side Hustle Tax Estimator — Premium Excel Template
Product for Gumroad ($12.99)

3 Sheets:
  1. Tax Estimator — inputs, tax calculations, quarterly payments
  2. Expense Tracker — monthly expense log (12 months × categories) with annual totals
  3. How To Use — instructions

Design: Forest Green palette (#1E8449 primary, #196F3D dark, #D5F5E3 input)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# DESIGN SYSTEM — Forest Green
# ============================================================
GREEN = "1E8449"
DARK_GREEN = "196F3D"
WHITE = "FFFFFF"
INPUT_GREEN = "D5F5E3"
LIGHT_GRAY = "F5F6FA"
MED_GRAY = "D5D8DC"
DARK_GRAY = "5D6D7E"
LIGHT_GREEN = "E8F8F5"
ACCENT = "27AE60"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, color="A9DFBF", italic=True)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=11, color="2C3E50")
FONT_INPUT = Font(name="Calibri", size=12, color=DARK_GREEN, bold=True)
FONT_VALUE = Font(name="Calibri", size=11, color="2C3E50")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=DARK_GREEN)
FONT_SMALL = Font(name="Calibri", size=9, color=DARK_GRAY, italic=True)
FONT_CTA = Font(name="Calibri", size=12, bold=True, color=DARK_GREEN)

FILL_GREEN = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_GREEN, end_color=INPUT_GREEN, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

THIN = Border(
    left=Side(style="thin", color=MED_GRAY), right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY), bottom=Side(style="thin", color=MED_GRAY),
)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def header_bar(ws, row, c1, c2, text, fill=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = FONT_SECTION
    cell.fill = fill or FILL_GREEN
    cell.alignment = ALIGN_C
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill or FILL_GREEN
        ws.cell(row=row, column=c).border = THIN


def label_input(ws, row, lc, vc, label, value=None, fmt=None):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=value)
    cv.font = FONT_INPUT
    cv.fill = FILL_INPUT
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def label_calc(ws, row, lc, vc, label, formula, fmt=None, bold=False):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=formula)
    cv.font = FONT_BOLD if bold else FONT_VALUE
    cv.fill = FILL_WHITE
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def cols(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ============================================================
# SHEET 1: TAX ESTIMATOR
# ============================================================
def build_tax_estimator(ws):
    ws.title = "Tax Estimator"
    ws.sheet_properties.tabColor = GREEN
    cols(ws, {"A": 2, "B": 36, "C": 18, "D": 4, "E": 36, "F": 18, "G": 2})

    for r in range(1, 80):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 7):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:F1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:F2")
    ws.row_dimensions[2].height = 38
    title = ws.cell(row=2, column=2, value="SIDE HUSTLE TAX ESTIMATOR")
    title.font = FONT_TITLE
    title.alignment = ALIGN_C
    ws.merge_cells("B3:F3")
    ws.row_dimensions[3].height = 22
    sub = ws.cell(row=3, column=2, value="Enter your numbers in the green cells. Tax calculations update automatically.")
    sub.font = FONT_SUBTITLE
    sub.alignment = ALIGN_C

    # ===== LEFT: INPUTS =====
    header_bar(ws, 5, 2, 3, "INCOME")
    label_input(ws, 6, 2, 3, "W-2 Salary ($)", 75000, "$#,##0")
    label_input(ws, 7, 2, 3, "Side Hustle Gross Income ($)", 25000, "$#,##0")

    header_bar(ws, 9, 2, 3, "SIDE HUSTLE EXPENSES")
    label_input(ws, 10, 2, 3, "Supplies/Materials ($)", 2000, "$#,##0")
    label_input(ws, 11, 2, 3, "Software/Tools ($)", 500, "$#,##0")
    label_input(ws, 12, 2, 3, "Home Office ($)", 1500, "$#,##0")
    label_input(ws, 13, 2, 3, "Vehicle/Mileage ($)", 1000, "$#,##0")
    label_input(ws, 14, 2, 3, "Marketing/Advertising ($)", 500, "$#,##0")
    label_input(ws, 15, 2, 3, "Other Deductible ($)", 500, "$#,##0")

    header_bar(ws, 17, 2, 3, "OTHER")
    label_input(ws, 18, 2, 3, "Filing: 1=Single 2=MFJ 3=HOH", 1, "0")
    label_input(ws, 19, 2, 3, "Standard Deduction? (1=Yes)", 1, "0")
    label_input(ws, 20, 2, 3, "State Tax Rate (e.g. 0.093)", 0.093, "0.0%")
    label_input(ws, 21, 2, 3, "Quarterly Payments Made ($)", 0, "$#,##0")

    # ===== RIGHT: CALCULATIONS =====
    header_bar(ws, 5, 5, 6, "TAX CALCULATIONS", FILL_DARK)

    label_calc(ws, 6, 5, 6, "Total Gross Income", "=C6+C7", "$#,##0")
    label_calc(ws, 7, 5, 6, "Total Expenses", "=C10+C11+C12+C13+C14+C15", "$#,##0")
    label_calc(ws, 8, 5, 6, "Side Hustle Net Profit", "=C7-F7", "$#,##0", bold=True)
    label_calc(ws, 9, 5, 6, "SE Taxable (92.35%)", "=F8*0.9235", "$#,##0")
    label_calc(ws, 10, 5, 6, "Self-Employment Tax (15.3%)", "=F9*0.153", "$#,##0")
    label_calc(ws, 11, 5, 6, "SE Tax Deduction (50%)", "=F10*0.5", "$#,##0")
    label_calc(ws, 12, 5, 6, "AGI", "=F6-F11", "$#,##0")
    label_calc(ws, 13, 5, 6, "Standard Deduction", "=IF(C19=1,IF(C18=1,16100,IF(C18=2,32200,24150)),0)", "$#,##0")
    label_calc(ws, 14, 5, 6, "Taxable Income", "=MAX(0,F12-F13)", "$#,##0", bold=True)

    header_bar(ws, 16, 5, 6, "FEDERAL TAX (2026 Brackets)")
    # Simplified: use effective rate approximation for Excel (full bracket calc would need many cells)
    ws.cell(row=17, column=5, value="Federal Tax (est.)").font = FONT_LABEL
    ws.cell(row=17, column=5).fill = FILL_GRAY
    ws.cell(row=17, column=5).border = THIN
    ws.cell(row=17, column=5).alignment = ALIGN_L
    ws.cell(row=17, column=6, value="=F14*0.22").font = FONT_VALUE
    ws.cell(row=17, column=6).fill = FILL_WHITE
    ws.cell(row=17, column=6).number_format = "$#,##0"
    ws.cell(row=17, column=6).border = THIN
    ws.cell(row=17, column=6).alignment = ALIGN_R

    label_calc(ws, 18, 5, 6, "State Tax", "=F14*C20", "$#,##0")
    label_calc(ws, 19, 5, 6, "Total Tax Liability", "=F10+F17+F18", "$#,##0", bold=True)
    label_calc(ws, 20, 5, 6, "Less: Quarterly Paid", "=-C21", "$#,##0")
    label_calc(ws, 21, 5, 6, "Remaining Tax Due", "=F19+F20", "$#,##0")
    label_calc(ws, 22, 5, 6, "Quarterly Payment (÷4)", "=F21/4", "$#,##0", bold=True)

    header_bar(ws, 24, 5, 6, "QUARTERLY SCHEDULE")
    for i, (qtr, due) in enumerate([("Q1", "Apr 15"), ("Q2", "Jun 15"), ("Q3", "Sep 15"), ("Q4", "Jan 15")]):
        ws.cell(row=25 + i, column=5, value=f"{qtr} — {due}").font = FONT_LABEL
        ws.cell(row=25 + i, column=5).fill = FILL_GRAY
        ws.cell(row=25 + i, column=5).border = THIN
        ws.cell(row=25 + i, column=6, value="=F22").font = FONT_VALUE
        ws.cell(row=25 + i, column=6).fill = FILL_WHITE
        ws.cell(row=25 + i, column=6).number_format = "$#,##0"
        ws.cell(row=25 + i, column=6).border = THIN

    ws.protection.sheet = True
    input_cells = [(6, 3), (7, 3), (10, 3), (11, 3), (12, 3), (13, 3), (14, 3), (15, 3), (18, 3), (19, 3), (20, 3), (21, 3)]
    for r, c in input_cells:
        ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 2: EXPENSE TRACKER
# ============================================================
def build_expense_tracker(wb):
    ws = wb.create_sheet("Expense Tracker")
    ws.sheet_properties.tabColor = ACCENT
    cols(ws, {"A": 2, "B": 14, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12, "J": 12, "K": 12, "L": 12, "M": 12, "N": 14})

    for r in range(1, 25):
        for c in range(1, 15):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    ws.merge_cells("B1:N2")
    ws.cell(row=1, column=2, value="EXPENSE TRACKER — Monthly Log").font = FONT_TITLE
    ws.cell(row=1, column=2).fill = FILL_DARK
    ws.cell(row=1, column=2).alignment = ALIGN_C
    for r in range(1, 3):
        for c in range(2, 15):
            ws.cell(row=r, column=c).fill = FILL_DARK

    categories = ["Supplies", "Software", "Home Office", "Vehicle", "Marketing", "Other"]
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    # Headers
    ws.cell(row=4, column=2, value="Category").font = FONT_HEADER
    ws.cell(row=4, column=2).fill = FILL_GREEN
    ws.cell(row=4, column=2).border = THIN
    ws.cell(row=4, column=2).alignment = ALIGN_C
    for j, m in enumerate(months):
        cell = ws.cell(row=4, column=3 + j, value=m)
        cell.font = FONT_HEADER
        cell.fill = FILL_GREEN
        cell.border = THIN
        cell.alignment = ALIGN_C
    ws.cell(row=4, column=15, value="Annual").font = FONT_HEADER
    ws.cell(row=4, column=15).fill = FILL_DARK
    ws.cell(row=4, column=15).border = THIN
    ws.cell(row=4, column=15).alignment = ALIGN_C

    for i, cat in enumerate(categories):
        r = 5 + i
        ws.cell(row=r, column=2, value=cat).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN
        for j in range(12):
            c = 3 + j
            cell = ws.cell(row=r, column=c, value=0)
            cell.font = FONT_INPUT
            cell.fill = FILL_INPUT
            cell.border = THIN
            cell.number_format = "$#,##0"
            cell.alignment = ALIGN_R
        ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").font = FONT_BOLD  # C:N = 12 months
        ws.cell(row=r, column=15).fill = FILL_LIGHT
        ws.cell(row=r, column=15).border = THIN
        ws.cell(row=r, column=15).number_format = "$#,##0"
        ws.cell(row=r, column=15).alignment = ALIGN_R

    # Total row
    r = 11
    ws.cell(row=r, column=2, value="Total").font = FONT_HEADER
    ws.cell(row=r, column=2).fill = FILL_DARK
    ws.cell(row=r, column=2).border = THIN
    for j in range(12):
        c = 3 + j
        col_letter = get_column_letter(c)
        ws.cell(row=r, column=c, value=f"=SUM({col_letter}5:{col_letter}10)").font = FONT_BOLD
        ws.cell(row=r, column=c).fill = FILL_LIGHT
        ws.cell(row=r, column=c).border = THIN
        ws.cell(row=r, column=c).number_format = "$#,##0"
    ws.cell(row=r, column=15, value="=SUM(O5:O10)").font = FONT_BOLD
    ws.cell(row=r, column=15).fill = FILL_DARK
    ws.cell(row=r, column=15).border = THIN
    ws.cell(row=r, column=15).number_format = "$#,##0"

    ws.protection.sheet = True
    for r in range(5, 11):
        for c in range(3, 15):
            ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 3: HOW TO USE
# ============================================================
def build_instructions(wb):
    ws = wb.create_sheet("How To Use")
    ws.sheet_properties.tabColor = DARK_GRAY
    cols(ws, {"A": 3, "B": 90})

    ws.merge_cells("A1:B2")
    c = ws.cell(row=1, column=1, value="HOW TO USE THE SIDE HUSTLE TAX ESTIMATOR")
    c.font = FONT_TITLE
    c.fill = FILL_DARK
    c.alignment = ALIGN_C
    for r in range(1, 3):
        for co in range(1, 3):
            ws.cell(row=r, column=co).fill = FILL_DARK

    sections = [
        ("QUICK START", [
            "1. Open the 'Tax Estimator' tab and enter your numbers in the GREEN cells",
            "2. W-2 salary, side hustle gross income, and deductible expenses",
            "3. Set Standard Deduction (1=Yes, 0=No for itemized)",
            "4. Enter your state tax rate (e.g., 0.093 for California 9.3%)",
            "5. Results update automatically: total tax, quarterly payments",
            "6. Use 'Expense Tracker' to log monthly expenses by category",
        ]),
        ("INPUT EXPLANATIONS", [
            "W-2 Salary: Your day job income (before taxes)",
            "Side Hustle Gross: Total revenue from freelance, Etsy, Uber, etc.",
            "Expenses: Deductible business expenses (supplies, software, home office, vehicle, marketing)",
            "Home Office: Simplified method ($5/sq ft) or actual expenses",
            "Standard Deduction: 2026 Single $16,100 | MFJ $32,200 | HOH $24,150",
            "State Tax Rate: Check your state's income tax rate. No tax states = 0",
        ]),
        ("TAX CALCULATIONS", [
            "Self-Employment Tax: 15.3% on 92.35% of net profit (SS + Medicare)",
            "SE Tax Deduction: Half of SE tax reduces your AGI",
            "Federal Tax: Uses 2026 brackets (10/12/22/24/32/35/37%)",
            "Quarterly Payments: Due Apr 15, Jun 15, Sep 15, Jan 15",
            "The Excel uses a simplified effective rate; the web app has full bracket logic",
        ]),
        ("EXPENSE TRACKER", [
            "Log monthly expenses by category (Supplies, Software, Home Office, etc.)",
            "Annual totals feed into your tax planning",
            "Keep receipts and records for IRS documentation",
        ]),
        ("IMPORTANT NOTES", [
            "This is an estimator only — consult a CPA for your specific situation",
            "State rates vary; use your state's effective or top rate",
            "SS wage base 2026: $184,500 (cap on Social Security portion of SE tax)",
            "© 2026 ClearMetric. For educational use only. Not financial or tax advice.",
        ]),
    ]

    r = 4
    for title, items in sections:
        ws.cell(row=r, column=2, value=title).font = Font(name="Calibri", size=12, bold=True, color=DARK_GREEN)
        ws.cell(row=r, column=2).fill = FILL_LIGHT
        ws.cell(row=r, column=2).border = THIN
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item).font = Font(name="Calibri", size=11, color="2C3E50")
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    print("Building Tax Estimator sheet...")
    build_tax_estimator(ws)

    print("Building Expense Tracker sheet...")
    build_expense_tracker(wb)

    print("Building How To Use sheet...")
    build_instructions(wb)

    wb.active = 0

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "output", "ClearMetric-Side-Hustle-Tax-Estimator.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Size: {os.path.getsize(out) / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
